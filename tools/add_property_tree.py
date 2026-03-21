"""Add a propertyTree worksheet to an existing Excel workbook.

Generates a tree view matching the format in cdifCore_properties.xlsx:
- Alternating entity/property columns
- Single-type values on same row as property
- Multiple-type values on subsequent rows
- Arrays shown with [] prefix
- Object references shown as '@id' with string type
"""
import openpyxl
from openpyxl.styles import Font, Alignment
import yaml
import os
import sys


def load_yaml(path):
    with open(path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def resolve_ref(ref, schema, schema_dir):
    """Resolve a $ref, returning (schema, dir, name)."""
    if ref.startswith('#/$defs/'):
        name = ref[len('#/$defs/'):]
        defn = schema.get('$defs', {}).get(name, {})
        if '$ref' in defn:
            return resolve_ref(defn['$ref'], schema, schema_dir)
        return defn, schema_dir, name
    elif ref.startswith(('http://', 'https://')):
        return None, None, ref.split('/')[-2]
    else:
        frag = None
        if '#' in ref:
            ref, frag = ref.split('#', 1)
        path = os.path.normpath(os.path.join(schema_dir, ref))
        if os.path.exists(path):
            resolved = load_yaml(path)
            rdir = os.path.dirname(path)
            name = os.path.basename(os.path.dirname(path))
            if frag and frag.startswith('/$defs/'):
                dname = frag[len('/$defs/'):]
                resolved = resolved.get('$defs', {}).get(dname, resolved)
                name = dname
            return resolved, rdir, name
        return None, None, ref


# Types to expand (show sub-properties)
EXPAND_TYPES = {
    'Identifier', 'identifier', 'LabeledLink', 'labeledLink',
    'DataDownload', 'dataDownload', 'WebAPI', 'webAPI',
    'CdifCatalogRecord', 'cdifCatalogRecord',
    'variableMeasured', 'VariableMeasured',
    'statisticalVariable', 'StatisticalVariable',
    'spatialExtent', 'SpatialExtent',
    'temporalExtent', 'TemporalExtent',
    'qualityMeasure', 'QualityMeasure',
}

# Types to show as just "object" (not expanded)
NO_EXPAND_TYPES = {
    'Person', 'person', 'Organization', 'organization',
    'DefinedTerm', 'definedTerm', 'Action', 'action',
    'Contributor', 'agentInRole', 'Funder', 'funder',
    'GeneratedBy', 'generatedBy', 'DerivedFrom', 'derivedFrom',
}


class TreeBuilder:
    def __init__(self):
        self.rows = []

    def add_row(self, **kwargs):
        self.rows.append(kwargs)

    def should_expand(self, name, depth):
        if depth >= 2:
            return False
        if name in EXPAND_TYPES:
            return True
        if name in NO_EXPAND_TYPES:
            return False
        return depth < 1

    def resolve_def(self, ref, parent_schema, schema_dir, aug_defs):
        if ref.startswith('#/$defs/'):
            name = ref[len('#/$defs/'):]
            if name in aug_defs:
                val = aug_defs[name]
                if isinstance(val, tuple):
                    return val[0], val[1], name
                return val, schema_dir, name
            defn = parent_schema.get('$defs', {}).get(name, {})
            if '$ref' in defn:
                return resolve_ref(defn['$ref'], parent_schema, schema_dir)
            return defn, schema_dir, name
        return resolve_ref(ref, parent_schema, schema_dir)

    def build_from_schema(self, schema, schema_dir):
        # Merge all properties from allOf composing BBs
        all_props = {}
        aug_defs = {}

        for entry in schema.get('allOf', []):
            if isinstance(entry, dict) and '$ref' in entry:
                ref = entry['$ref']
                if not ref.startswith('#'):
                    resolved, rdir, _ = resolve_ref(ref, schema, schema_dir)
                    if resolved:
                        for k, v in resolved.get('properties', {}).items():
                            if k not in all_props:
                                all_props[k] = (v, rdir, resolved)
                        for dn, ds in resolved.get('$defs', {}).items():
                            if dn not in aug_defs:
                                if isinstance(ds, dict) and '$ref' in ds:
                                    rs, rd, _ = resolve_ref(ds['$ref'], resolved, rdir)
                                    if rs:
                                        aug_defs[dn] = (rs, rd)
                                        continue
                                aug_defs[dn] = (ds, rdir)
                        for sub in resolved.get('allOf', []):
                            if isinstance(sub, dict) and sub.get('properties'):
                                for k, v in sub['properties'].items():
                                    if k not in all_props:
                                        all_props[k] = (v, rdir, resolved)
            elif isinstance(entry, dict) and entry.get('properties'):
                for k, v in entry['properties'].items():
                    if k not in all_props:
                        all_props[k] = (v, schema_dir, schema)

        for k, v in schema.get('properties', {}).items():
            if k not in all_props:
                all_props[k] = (v, schema_dir, schema)

        # Also merge profile-level $defs
        for dn, ds in schema.get('$defs', {}).items():
            if dn not in aug_defs:
                if isinstance(ds, dict) and '$ref' in ds:
                    rs, rd, _ = resolve_ref(ds['$ref'], schema, schema_dir)
                    if rs:
                        aug_defs[dn] = (rs, rd)
                        continue
                aug_defs[dn] = (ds, schema_dir)

        # Headers
        self.add_row(c1='object', c2='property', c3='object/data type',
                     c4='property', c5='object/data type',
                     c6='property', c7='object/datatype')

        self.add_row(c1='Schema:Dataset')

        for prop_name, (prop_schema, pdir, parent_schema) in all_props.items():
            self.emit_property(prop_name, prop_schema, 2, pdir, parent_schema, aug_defs, 0)

    def emit_property(self, name, schema, prop_col, schema_dir, parent_schema, aug_defs, depth):
        if not isinstance(schema, dict):
            row = {f'c{prop_col}': name}
            if schema is not None:
                row[f'c{prop_col+1}'] = str(schema)
            self.add_row(**row)
            return

        type_col = prop_col + 1

        # Skip @context expansion
        if name == '@context':
            self.add_row(**{f'c{prop_col}': name})
            return

        # Direct $ref
        if '$ref' in schema:
            resolved, rdir, rname = self.resolve_def(schema['$ref'], parent_schema, schema_dir, aug_defs)
            if resolved and self.should_expand(rname, depth):
                self.add_row(**{f'c{prop_col}': name})
                self.add_row(**{f'c{type_col}': rname})
                self.emit_entity_props(resolved, type_col, rdir or schema_dir, aug_defs, depth + 1)
            elif rname in NO_EXPAND_TYPES:
                self.add_row(**{f'c{prop_col}': name, f'c{type_col}': rname,
                              f'c{min(type_col+2, 7)}': 'object'})
            else:
                self.add_row(**{f'c{prop_col}': name, f'c{type_col}': rname or 'object'})
            return

        t = schema.get('type')

        # Simple scalar types
        if t in ('string', 'number', 'integer', 'boolean'):
            fmt = schema.get('format', '')
            ts = f'{t}({fmt})' if fmt else t
            self.add_row(**{f'c{prop_col}': name, f'c{type_col}': ts})
            return

        # Multi-type: type: [string, number]
        if isinstance(t, list):
            self.add_row(**{f'c{prop_col}': name, f'c{type_col}': ' | '.join(str(x) for x in t)})
            return

        # Array type
        if t == 'array':
            items = schema.get('items', {})
            if isinstance(items, dict):
                if '$ref' in items:
                    resolved, rdir, rname = self.resolve_def(items['$ref'], parent_schema, schema_dir, aug_defs)
                    if resolved and self.should_expand(rname, depth):
                        self.add_row(**{f'c{prop_col}': name})
                        self.add_row(**{f'c{type_col}': f'[{rname}]'})
                        self.emit_entity_props(resolved, type_col, rdir or schema_dir, aug_defs, depth + 1)
                    elif rname in NO_EXPAND_TYPES:
                        self.add_row(**{f'c{prop_col}': name, f'c{type_col}': f'[{rname}]',
                                      f'c{min(type_col+2, 7)}': 'object'})
                    else:
                        self.add_row(**{f'c{prop_col}': name, f'c{type_col}': f'[{rname}]'})
                    return

                it = items.get('type')
                if it in ('string', 'number', 'integer', 'boolean'):
                    fmt = items.get('format', '')
                    ts = f'{it}({fmt})' if fmt else it
                    self.add_row(**{f'c{prop_col}': name, f'c{type_col}': f'[{ts}]'})
                    return

                if 'anyOf' in items:
                    self.add_row(**{f'c{prop_col}': name})
                    self.emit_anyof(items['anyOf'], type_col, schema_dir, parent_schema, aug_defs, depth, is_array=True)
                    return

                if it == 'object' or items.get('properties'):
                    self.add_row(**{f'c{prop_col}': name, f'c{type_col}': '[object]'})
                    if items.get('properties') and depth < 2:
                        self.emit_obj_props(items, type_col, schema_dir, parent_schema, aug_defs, depth + 1)
                    return

            self.add_row(**{f'c{prop_col}': name, f'c{type_col}': '[any]'})
            return

        # anyOf at property level
        if 'anyOf' in schema:
            self.add_row(**{f'c{prop_col}': name})
            self.emit_anyof(schema['anyOf'], type_col, schema_dir, parent_schema, aug_defs, depth)
            return

        # Object with properties
        if t == 'object' or schema.get('properties'):
            props = schema.get('properties', {})
            # Simple @id reference
            if list(props.keys()) == ['@id']:
                self.add_row(**{f'c{prop_col}': name, f'c{type_col}': 'object reference',
                              f'c{type_col+1}': '@id', f'c{type_col+2}': 'string'})
                return
            # @list pattern
            if '@list' in props:
                list_schema = props['@list']
                if list_schema.get('type') == 'array' and 'items' in list_schema:
                    self.add_row(**{f'c{prop_col}': name})
                    items = list_schema['items']
                    if 'anyOf' in items:
                        self.emit_anyof(items['anyOf'], type_col, schema_dir, parent_schema, aug_defs, depth, is_array=True, prefix='@list ')
                    return
            # Generic object
            self.add_row(**{f'c{prop_col}': name, f'c{type_col}': 'object'})
            if depth < 2 and props:
                self.emit_obj_props(schema, type_col, schema_dir, parent_schema, aug_defs, depth + 1)
            return

        # Fallback
        self.add_row(**{f'c{prop_col}': name})

    def emit_anyof(self, options, type_col, schema_dir, parent_schema, aug_defs, depth, is_array=False, prefix=''):
        lb = '[' if is_array else ''
        rb = ']' if is_array else ''

        for opt in options:
            if not isinstance(opt, dict):
                continue
            if '$ref' in opt:
                resolved, rdir, rname = self.resolve_def(opt['$ref'], parent_schema, schema_dir, aug_defs)
                if resolved and self.should_expand(rname, depth):
                    self.add_row(**{f'c{type_col}': f'{lb}{prefix}{rname}{rb}'})
                    self.emit_entity_props(resolved, type_col, rdir or schema_dir, aug_defs, depth + 1)
                elif rname in NO_EXPAND_TYPES:
                    self.add_row(**{f'c{type_col}': f'{lb}{prefix}{rname}{rb}',
                                  f'c{min(type_col+2, 7)}': 'object'})
                else:
                    self.add_row(**{f'c{type_col}': f'{lb}{prefix}{rname}{rb}'})
                continue

            ot = opt.get('type')
            if ot in ('string', 'number', 'integer', 'boolean'):
                fmt = opt.get('format', '')
                ts = f'{ot}({fmt})' if fmt else ot
                self.add_row(**{f'c{type_col}': f'{lb}{prefix}{ts}{rb}'})
                continue

            if ot == 'array' and 'items' in opt:
                items = opt['items']
                if 'anyOf' in items:
                    self.emit_anyof(items['anyOf'], type_col, schema_dir, parent_schema, aug_defs, depth, is_array=True)
                elif '$ref' in items:
                    _, _, rn = self.resolve_def(items['$ref'], parent_schema, schema_dir, aug_defs)
                    self.add_row(**{f'c{type_col}': f'[{prefix}{rn}]'})
                elif items.get('type'):
                    self.add_row(**{f'c{type_col}': f'[{prefix}{items["type"]}]'})
                continue

            if ot == 'object' or opt.get('properties'):
                props = opt.get('properties', {})
                if list(props.keys()) == ['@id']:
                    self.add_row(**{f'c{type_col}': f'{lb}{prefix}object reference{rb}',
                                  f'c{type_col+1}': '@id', f'c{type_col+2}': 'string'})
                else:
                    self.add_row(**{f'c{type_col}': f'{lb}{prefix}object{rb}'})
                continue

    def emit_entity_props(self, schema, type_col, schema_dir, aug_defs, depth):
        if depth > 3:
            return
        prop_col = type_col + 1
        if prop_col > 7:
            return
        for pname, pschema in schema.get('properties', {}).items():
            if pname == '@context':
                continue
            self.emit_property(pname, pschema, prop_col, schema_dir, schema, aug_defs, depth)

    def emit_obj_props(self, schema, parent_type_col, schema_dir, parent_schema, aug_defs, depth):
        prop_col = parent_type_col + 1
        if prop_col > 7:
            return
        for pname, pschema in schema.get('properties', {}).items():
            if pname == '@context':
                continue
            self.emit_property(pname, pschema, prop_col, schema_dir, parent_schema, aug_defs, depth)

    def write_to_workbook(self, wb_path):
        wb = openpyxl.load_workbook(wb_path)

        if 'propertyTree' in wb.sheetnames:
            del wb['propertyTree']

        ws = wb.create_sheet('propertyTree')

        for row_idx, row_data in enumerate(self.rows, 1):
            for key, value in row_data.items():
                col = int(key[1:])
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = Font(size=11)
                cell.alignment = Alignment(vertical='top')

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 28
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 28
        ws.column_dimensions['G'].width = 20

        wb.save(wb_path)
        print(f'Added propertyTree ({len(self.rows)} rows) to {os.path.basename(wb_path)}')


def main():
    BB_ROOT = r'C:\Users\smrTu\OneDrive\Documents\GithubC\CDIF\metadataBuildingBlocksFork\_sources'

    targets = [
        (os.path.join(BB_ROOT, 'profiles', 'cdifProfiles', 'CDIFDiscoveryProfile', 'schema.yaml'),
         os.path.join(BB_ROOT, 'profiles', 'cdifProfiles', 'CDIFDiscoveryProfile', 'CDIFDiscoveryProfile_properties.xlsx')),
        (os.path.join(BB_ROOT, 'profiles', 'cdifProfiles', 'CDIFDataDescriptionProfile', 'schema.yaml'),
         os.path.join(BB_ROOT, 'profiles', 'cdifProfiles', 'CDIFDataDescriptionProfile', 'CDIFDataDescriptionProfile_properties.xlsx')),
    ]

    for schema_path, wb_path in targets:
        print(f'\nProcessing: {os.path.basename(schema_path)}')
        schema = load_yaml(schema_path)
        schema_dir = os.path.dirname(schema_path)

        builder = TreeBuilder()
        builder.build_from_schema(schema, schema_dir)
        builder.write_to_workbook(wb_path)


if __name__ == '__main__':
    main()
