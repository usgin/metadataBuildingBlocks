#!/usr/bin/env python3
"""
Generate propertyTree_2 worksheets from resolved JSON Schemas.

Walks a resolved schema and produces a spreadsheet tree following the CDIF
property-tree convention:

    col 0: root object type
    col 1: property (depth 0)
    col 2: options  (depth 0)
    col 3: property (depth 1)
    col 4: options  (depth 1)
    ...

Suffix conventions:
    -- string               literal string
    -- string(uri)          string with URI format
    -- string(date)         string with date format
    -- boolean              boolean
    -- number               number
    -- object reference     JSON-LD {@id: ...} reference
    -- object               nested object
    -- CHOICE               anyOf with mixed types
    [...]                   square brackets = array (0..* or 1..*)
"""

import json
import os
import sys
import shutil
import tempfile
import argparse

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

MAX_DEPTH = 6
MAX_COLS = 2 + 2 * MAX_DEPTH  # object col + depth levels * 2


# ── helpers ──────────────────────────────────────────────────────────────

def get_type_contains(schema):
    """Return the @type contains value, if any."""
    tp = schema.get("properties", {}).get("@type", {})
    c = tp.get("contains", {})
    return c.get("const")


def is_id_only_object(schema):
    """True if this object schema is just {@id: string}."""
    if schema.get("type") != "object":
        return False
    props = set(schema.get("properties", {}).keys())
    return props <= {"@id"}


def is_lang_tagged_pattern(any_of):
    """Detect the string-or-language-tagged-value pattern."""
    has_string = False
    has_lang = False
    for item in any_of:
        if item.get("type") == "string":
            has_string = True
        elif item.get("type") == "object":
            pkeys = set(item.get("properties", {}).keys())
            if "@value" in pkeys and "@language" in pkeys:
                has_lang = True
        elif item.get("type") == "array":
            sub = item.get("items", {})
            if sub.get("type") == "object":
                pkeys = set(sub.get("properties", {}).keys())
                if "@value" in pkeys and "@language" in pkeys:
                    has_lang = True
            for s in sub.get("anyOf", []):
                if s.get("type") == "object":
                    pkeys = set(s.get("properties", {}).keys())
                    if "@value" in pkeys and "@language" in pkeys:
                        has_lang = True
    return has_string and has_lang


def classify_anyof(any_of):
    """Return a list of (kind, schema[, type_contains]) tuples."""
    options = []
    for item in any_of:
        t = item.get("type")
        if t == "string":
            if "enum" in item:
                options.append(("enum", item))
            else:
                fmt = item.get("format", "")
                options.append(("string", item))
        elif t in ("number", "integer"):
            options.append(("number", item))
        elif t == "boolean":
            options.append(("boolean", item))
        elif t == "array":
            options.append(("array", item))
        elif t == "object":
            if is_id_only_object(item):
                options.append(("object_reference", item))
            else:
                tc = get_type_contains(item)
                if tc:
                    options.append(("typed_object", item, tc))
                else:
                    options.append(("object", item))
        elif "enum" in item:
            options.append(("enum", item))
        else:
            options.append(("unknown", item))
    return options


def make_row():
    return [None] * MAX_COLS


def string_suffix(schema):
    fmt = schema.get("format", "")
    if fmt in ("uri", "iri", "uri-reference", "iri-reference"):
        return "string(uri)"
    if fmt == "date":
        return "string(date)"
    if fmt == "date-time":
        return "string(date-time)"
    return "string"


# ── main tree walker ─────────────────────────────────────────────────────

def emit_property(prop_name, prop_schema, depth, rows, visited_types):
    """Emit row(s) for one property."""
    if depth > MAX_DEPTH:
        return
    prop_col = 1 + 2 * depth
    opts_col = 2 + 2 * depth
    if prop_col >= MAX_COLS:
        return

    is_array = prop_schema.get("type") == "array"

    # ── @context ──
    if prop_name == "@context":
        row = make_row()
        row[prop_col] = "@context -- standard context"
        rows.append(row)
        return

    # ── @id ──
    if prop_name == "@id":
        row = make_row()
        row[prop_col] = "@id -- string(uri)"
        rows.append(row)
        return

    # ── @type ──
    if prop_name == "@type":
        contains = None
        c = prop_schema.get("contains", {})
        contains = c.get("const")
        row = make_row()
        if contains:
            row[prop_col] = f"@type -- [string](contains {contains})"
        else:
            row[prop_col] = "@type -- [string]"
        rows.append(row)
        return

    # ── effective schema (unwrap array) ──
    effective = prop_schema.get("items", {}) if is_array else prop_schema

    # ── type as array (e.g. ["string", "number"]) → CHOICE ──
    ptype = effective.get("type")
    if isinstance(ptype, list):
        bracket = "[CHOICE]" if is_array else "CHOICE"
        row = make_row()
        row[prop_col] = f"{prop_name} -- {bracket}"
        rows.append(row)
        for t in ptype:
            r = make_row()
            r[opts_col] = t
            rows.append(r)
        return

    # ── anyOf / oneOf ──
    any_of = effective.get("anyOf", effective.get("oneOf"))
    # If anyOf items are only constraint refinements (e.g. different 'required'
    # sets) and the schema already has a concrete 'type', skip the anyOf and
    # treat as the base type.
    if any_of and all(set(item.keys()) <= {"required", "if", "then", "else",
                                            "not", "description"}
                      for item in any_of):
        any_of = None

    if any_of:
        # language-tagged pattern → treat as simple string
        if is_lang_tagged_pattern(any_of):
            sfx = "[string]" if is_array else "string"
            row = make_row()
            row[prop_col] = f"{prop_name} -- {sfx}"
            rows.append(row)
            return

        options = classify_anyof(any_of)
        kind_set = set(o[0] for o in options)

        # all-string → simple string
        if kind_set <= {"string"}:
            sfx = "[string]" if is_array else "string"
            row = make_row()
            row[prop_col] = f"{prop_name} -- {sfx}"
            rows.append(row)
            return

        # all object-reference → simple ref
        if kind_set <= {"object_reference"}:
            sfx = "[object reference]" if is_array else "object reference"
            row = make_row()
            row[prop_col] = f"{prop_name} -- {sfx}"
            rows.append(row)
            return

        # genuine CHOICE
        bracket = "[CHOICE]" if is_array else "CHOICE"
        row = make_row()
        row[prop_col] = f"{prop_name} -- {bracket}"
        rows.append(row)

        for opt in options:
            emit_choice_option(opt, opts_col, depth, rows, visited_types)
        return

    # ── simple types ──
    ptype = effective.get("type")

    if ptype == "string":
        sfx = string_suffix(effective)
        if is_array:
            sfx = f"[{sfx}]"
        row = make_row()
        row[prop_col] = f"{prop_name} -- {sfx}"
        rows.append(row)
        return

    if ptype == "boolean":
        sfx = "[boolean]" if is_array else "boolean"
        row = make_row()
        row[prop_col] = f"{prop_name} -- {sfx}"
        rows.append(row)
        return

    if ptype in ("number", "integer"):
        sfx = "[number]" if is_array else "number"
        row = make_row()
        row[prop_col] = f"{prop_name} -- {sfx}"
        rows.append(row)
        return

    if "enum" in effective:
        vals = ", ".join(f'"{v}"' for v in effective["enum"])
        sfx = f"[enum]" if is_array else "enum"
        row = make_row()
        row[prop_col] = f"{prop_name} -- {sfx} -- {vals}"
        rows.append(row)
        return

    if ptype == "object":
        emit_object_property(prop_name, effective, is_array, depth, rows, visited_types)
        return

    # ── const ──
    if "const" in effective:
        row = make_row()
        row[prop_col] = f'{prop_name} -- const("{effective["const"]}")'
        rows.append(row)
        return

    # fallback
    row = make_row()
    row[prop_col] = f"{prop_name}"
    rows.append(row)


def emit_object_property(prop_name, schema, is_array, depth, rows, visited_types):
    """Emit rows for a property whose value is an object."""
    prop_col = 1 + 2 * depth
    opts_col = 2 + 2 * depth

    if is_id_only_object(schema):
        sfx = "[object reference]" if is_array else "object reference"
        row = make_row()
        row[prop_col] = f"{prop_name} -- {sfx}"
        rows.append(row)
        return

    tc = get_type_contains(schema)
    sfx = "[object]" if is_array else "object"

    row = make_row()
    row[prop_col] = f"{prop_name} -- {sfx}"
    if tc and opts_col < MAX_COLS:
        row[opts_col] = f"@type -- [string](contains {tc})"
    rows.append(row)

    # recursion guard
    if tc and tc in visited_types:
        return
    new_visited = visited_types | ({tc} if tc else set())

    # emit child properties
    if depth + 1 <= MAX_DEPTH:
        child_props = {k: v for k, v in schema.get("properties", {}).items()
                       if k != "@type"}
        for cname, cschema in child_props.items():
            emit_property(cname, cschema, depth + 1, rows, new_visited)


def emit_choice_option(opt, opts_col, depth, rows, visited_types):
    """Emit one option within a CHOICE."""
    if opts_col >= MAX_COLS:
        return

    kind = opt[0]
    item = opt[1]

    if kind == "string":
        fmt = item.get("format", "")
        label = f"string({fmt})" if fmt else "string"
        row = make_row()
        row[opts_col] = label
        rows.append(row)

    elif kind == "number":
        row = make_row()
        row[opts_col] = "number"
        rows.append(row)

    elif kind == "boolean":
        row = make_row()
        row[opts_col] = "boolean"
        rows.append(row)

    elif kind == "object_reference":
        row = make_row()
        row[opts_col] = "object reference"
        rows.append(row)

    elif kind == "enum":
        vals = ", ".join(f'"{v}"' for v in item.get("enum", []))
        row = make_row()
        row[opts_col] = f"enum -- {vals}"
        rows.append(row)

    elif kind == "typed_object":
        tc = opt[2]
        row = make_row()
        row[opts_col] = f"@type -- [string](contains {tc})"
        rows.append(row)

        if tc in visited_types:
            return
        new_visited = visited_types | {tc}

        if depth + 1 <= MAX_DEPTH:
            child_props = {k: v for k, v in item.get("properties", {}).items()
                           if k != "@type"}
            for cname, cschema in child_props.items():
                emit_property(cname, cschema, depth + 1, rows, new_visited)

    elif kind == "object":
        # untyped object — show its properties
        row = make_row()
        row[opts_col] = "object"
        rows.append(row)

        if depth + 1 <= MAX_DEPTH:
            child_props = {k: v for k, v in item.get("properties", {}).items()
                           if k != "@type"}
            for cname, cschema in child_props.items():
                emit_property(cname, cschema, depth + 1, rows, visited_types)

    elif kind == "array":
        # array option — language tagged etc; just show "array"
        row = make_row()
        row[opts_col] = "array"
        rows.append(row)

    else:
        row = make_row()
        row[opts_col] = str(kind)
        rows.append(row)


# ── top-level entry ──────────────────────────────────────────────────────

def generate_property_tree(resolved_schema_path):
    """Generate the full property tree rows from a resolved schema."""
    with open(resolved_schema_path) as f:
        schema = json.load(f)

    # determine root type
    tc = get_type_contains(schema)
    root_label = tc if tc else "root"

    # header
    header = make_row()
    for i in range(0, MAX_COLS, 2):
        if i == 0:
            header[i] = "object"
        else:
            header[i] = "options"
    for i in range(1, MAX_COLS, 2):
        header[i] = "property"

    rows = [header]

    # root object row
    root_row = make_row()
    root_row[0] = root_label
    rows.append(root_row)

    # walk all properties
    props = schema.get("properties", {})
    for pname, pschema in props.items():
        emit_property(pname, pschema, 0, rows, set())

    return rows


def write_worksheet(ws, rows):
    """Write rows to an openpyxl worksheet with formatting."""
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                              fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    for r_idx, row_data in enumerate(rows, start=1):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = wrap
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill

    # auto-width (rough estimate)
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


# ── CLI ──────────────────────────────────────────────────────────────────

def add_sheet_to_existing(xlsx_path, sheet_name, rows):
    """Add a worksheet to an existing xlsx file."""
    tmp = os.path.join(tempfile.gettempdir(),
                       f"tmp_{os.path.basename(xlsx_path)}")
    shutil.copy2(xlsx_path, tmp)
    wb = openpyxl.load_workbook(tmp)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    write_worksheet(ws, rows)
    wb.save(xlsx_path)
    wb.close()
    os.remove(tmp)
    print(f"  Added '{sheet_name}' to {xlsx_path}")


def create_new_workbook(xlsx_path, sheet_name, rows):
    """Create a new xlsx file with one worksheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    write_worksheet(ws, rows)
    wb.save(xlsx_path)
    wb.close()
    print(f"  Created {xlsx_path} with sheet '{sheet_name}'")


def main():
    parser = argparse.ArgumentParser(
        description="Generate propertyTree_2 worksheets from resolved schemas")
    parser.add_argument("--profile", choices=["codelist", "discovery",
                                               "datadescription", "all"],
                        default="all")
    args = parser.parse_args()

    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    profiles_dir = os.path.join(base, "_sources", "profiles", "cdifProfiles")

    targets = []
    if args.profile in ("codelist", "all"):
        targets.append({
            "name": "CDIFCodelistProfile",
            "resolved": os.path.join(profiles_dir,
                                     "CDIFCodelistProfile", "resolvedSchema.json"),
            "xlsx": os.path.join(profiles_dir,
                                 "CDIFCodelistProfile",
                                 "CDIFCodelistProfile_properties.xlsx"),
            "existing": False,
        })
    if args.profile in ("discovery", "all"):
        targets.append({
            "name": "CDIFDiscoveryProfile",
            "resolved": os.path.join(profiles_dir,
                                     "CDIFDiscoveryProfile", "resolvedSchema.json"),
            "xlsx": os.path.join(profiles_dir,
                                 "CDIFDiscoveryProfile",
                                 "CDIFDiscovery_properties.xlsx"),
            "existing": True,
        })
    if args.profile in ("datadescription", "all"):
        targets.append({
            "name": "CDIFDataDescriptionProfile",
            "resolved": os.path.join(profiles_dir,
                                     "CDIFDataDescriptionProfile",
                                     "resolvedSchema.json"),
            "xlsx": os.path.join(profiles_dir,
                                 "CDIFDataDescriptionProfile",
                                 "CDIFDataDescriptionProfile_properties.xlsx"),
            "existing": True,
        })

    for t in targets:
        print(f"Generating propertyTree_2 for {t['name']}...")
        rows = generate_property_tree(t["resolved"])
        print(f"  {len(rows)} rows generated")

        if t["existing"] and os.path.exists(t["xlsx"]):
            add_sheet_to_existing(t["xlsx"], "propertyTree_2", rows)
        else:
            create_new_workbook(t["xlsx"], "propertyTree_2", rows)


if __name__ == "__main__":
    main()
